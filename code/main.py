import cv2
import easyocr
import time
import re
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

reader = easyocr.Reader(['en'])

cap = cv2.VideoCapture(4)

cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

plate_pattern = r'^[A-Z]{2}[0-9]{1,2}[A-Z]{1,3}[0-9]{4}$'

excel_file = "vehicle_log.xlsx"

if not os.path.exists(excel_file):

    wb = Workbook()
    ws = wb.active

    ws.append([
        "Plate Number",
        "Date",
        "Time",
        "Image Name"
    ])

    wb.save(excel_file)

last_plate = ""
last_saved = {}

last_ocr_time = 0

frame_count = 0
fps = 0
fps_start = time.time()


def open_gate():

    print("\n====================")
    print("GATE OPENED")
    print("====================")

    time.sleep(3)

    print("====================")
    print("GATE CLOSED")
    print("====================\n")


while True:

    ret, frame = cap.read()

    if not ret:
        break

    frame_count += 1

    elapsed = time.time() - fps_start

    if elapsed >= 1:

        fps = frame_count / elapsed

        frame_count = 0
        fps_start = time.time()

    h, w = frame.shape[:2]

    x1 = int(w * 0.25)
    y1 = int(h * 0.35)

    x2 = int(w * 0.75)
    y2 = int(h * 0.65)

    cv2.rectangle(
        frame,
        (x1, y1),
        (x2, y2),
        (0, 255, 0),
        2
    )

    cv2.putText(
        frame,
        "PLACE PLATE HERE",
        (x1, y1 - 10),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.7,
        (0, 255, 0),
        2
    )

    current_time = time.time()

    if current_time - last_ocr_time >= 3:

        roi = frame[y1:y2, x1:x2]

        cv2.imshow("OCR Input", roi)

        results = reader.readtext(roi)

        detected_plate = ""

        for r in results:

            text = r[1]

            text = text.replace(" ", "")
            text = text.replace("-", "")
            text = text.upper()

            if re.match(plate_pattern, text):

                detected_plate = text
                break

        last_plate = detected_plate

        if detected_plate != "":

            now = time.time()

            if detected_plate not in last_saved or now - last_saved[detected_plate] > 90:

                timestamp = datetime.now()

                date_str = timestamp.strftime("%d-%m-%Y")
                time_str = timestamp.strftime("%H:%M:%S")

                image_name = (
                    detected_plate +
                    "_" +
                    timestamp.strftime("%Y%m%d_%H%M%S") +
                    ".jpg"
                )

                image_path = os.path.join(
                    "captured_vehicles",
                    image_name
                )

                cv2.imwrite(image_path, frame)

                wb = load_workbook(excel_file)
                ws = wb.active

                ws.append([
                    detected_plate,
                    date_str,
                    time_str,
                    image_name
                ])

                wb.save(excel_file)

                last_saved[detected_plate] = now

                print(
                    "Saved:",
                    detected_plate,
                    date_str,
                    time_str
                )

                open_gate()

        last_ocr_time = current_time

    cv2.putText(
        frame,
        f"FPS: {fps:.1f}",
        (20, 40),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.8,
        (0, 255, 0),
        2
    )

    if last_plate != "":

        cv2.putText(
            frame,
            last_plate,
            (20, 80),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.8,
            (0, 255, 0),
            2
        )

    cv2.imshow(
        "Indian ANPR",
        frame
    )

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
